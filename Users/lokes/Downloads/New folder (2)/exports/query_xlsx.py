import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\lokes\Downloads\New folder (2)\alphaveto\enquiries\institutional_leads.xlsx', data_only=True)
rows = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
print(json.dumps(rows, indent=2, default=str))
print(f'\nTotal rows: {len(rows)}')
